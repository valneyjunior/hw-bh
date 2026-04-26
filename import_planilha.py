"""
Lê Cadastro BH2.xlsx e gera app/import.php com os dados embutidos.
Após rodar, acesse localhost:3000/import.php como admin para importar.
"""
import openpyxl
from datetime import datetime
import uuid
import os
import sys

XLSX   = r"C:\Users\valney.junior\OneDrive - hostweb.cloud\Documentos\Cadastro BH2.xlsx"
OUTPUT = os.path.join(os.path.dirname(__file__), "app", "import.php")

SHEETS = ["Hanny", "Danilo", "Hudson", "Filippe"]

print(f"Lendo {XLSX}...")
try:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
except FileNotFoundError:
    print("ERRO: Arquivo não encontrado. Feche o Excel e tente novamente.")
    sys.exit(1)

def new_id():
    return str(uuid.uuid4())

def php_str(s):
    if s is None:
        return "''"
    return "'" + str(s).replace("\\", "\\\\").replace("'", "\\'").strip() + "'"

def php_dt(v):
    if isinstance(v, datetime) and v.year >= 2020:
        return "'" + v.strftime("%Y-%m-%d %H:%M:%S") + "'"
    return None

users   = []
records = []

for sheet_name in SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"  Aba '{sheet_name}' não encontrada — ignorando.")
        continue

    ws  = wb[sheet_name]
    uid = new_id()
    users.append({"id": uid, "name": sheet_name})

    count = 0
    for row in ws.iter_rows(min_row=2, max_row=300, values_only=True):
        started_at = row[0] if len(row) > 0 else None
        ended_at   = row[1] if len(row) > 1 else None
        ticket     = row[2] if len(row) > 2 else None
        description= row[3] if len(row) > 3 else None
        status     = row[5] if len(row) > 5 else None

        s_dt = php_dt(started_at)
        e_dt = php_dt(ended_at)

        if not s_dt or not e_dt or not ticket:
            continue
        if isinstance(started_at, datetime) and isinstance(ended_at, datetime):
            if ended_at <= started_at:
                continue

        records.append({
            "id":          new_id(),
            "user_id":     uid,
            "started_at":  s_dt,
            "ended_at":    e_dt,
            "ticket":      php_str(ticket),
            "description": php_str(description),
            "validated":   bool(status),
        })
        count += 1

    print(f"  {sheet_name}: {count} registros")

print(f"\nTotal: {len(users)} colaboradores, {len(records)} registros")

# ── Gera PHP ─────────────────────────────────────────────────────────────────
lines = []
lines.append("<?php")
lines.append("require_once 'includes/auth.php';")
lines.append("require_once 'includes/db.php';")
lines.append("require_once 'includes/functions.php';")
lines.append("$admin = requireAdmin();")
lines.append("")
lines.append("$defaultPass = password_hash('Hostweb@2025', PASSWORD_BCRYPT, ['cost' => 12]);")
lines.append("$now         = date('Y-m-d H:i:s');")
lines.append("$db          = getDb();")
lines.append("")

# Users array
lines.append("$users = [")
for u in users:
    email = u["name"].lower() + "@hostweb.cloud"
    lines.append(f"  ['id'=>'{u['id']}','name'=>'{u['name']}','email'=>'{email}'],")
lines.append("];")
lines.append("")

# Records array
lines.append("$records = [")
for r in records:
    val = "true" if r["validated"] else "false"
    lines.append(
        f"  ['id'=>'{r['id']}','uid'=>'{r['user_id']}',"
        f"'s'=>{r['started_at']},'e'=>{r['ended_at']},"
        f"'t'=>{r['ticket']},'d'=>{r['description']},'v'=>{val}],"
    )
lines.append("];")
lines.append("")

# Logic
lines.append(r"""
$created = $skipped = $recs = 0;
$idMap   = [];

foreach ($users as $u) {
    $stmt = $db->prepare("SELECT id FROM users WHERE email = ?");
    $stmt->execute([$u['email']]);
    $existing = $stmt->fetch();
    if ($existing) {
        $idMap[$u['id']] = $existing['id'];
        $skipped++;
    } else {
        $db->prepare("INSERT INTO users (id,name,email,password_hash,role,must_change_pass) VALUES (?,?,?,?,'collaborator',1)")
           ->execute([$u['id'], $u['name'], $u['email'], $defaultPass]);
        $idMap[$u['id']] = $u['id'];
        $created++;
    }
}

foreach ($records as $r) {
    $uid = $idMap[$r['uid']] ?? $r['uid'];
    $vat = $r['v'] ? $now  : null;
    $vby = $r['v'] ? $admin['id'] : null;
    $db->prepare("INSERT IGNORE INTO records (id,user_id,started_at,ended_at,ticket,description,validated_at,validated_by)
                  VALUES (?,?,?,?,?,?,?,?)")
       ->execute([$r['id'], $uid, $r['s'], $r['e'], $r['t'], $r['d'], $vat, $vby]);
    $recs++;
}
?>
<!DOCTYPE html>
<html lang="pt-BR">
<head><meta charset="UTF-8"><title>Importação</title>
<script src="https://cdn.tailwindcss.com"></script></head>
<body class="bg-gray-50 min-h-screen flex items-center justify-center p-4">
<div class="bg-white rounded-2xl shadow-lg p-8 max-w-sm w-full text-center">
  <div class="w-16 h-16 bg-green-100 rounded-full flex items-center justify-center mx-auto mb-4">
    <svg class="w-8 h-8 text-green-600" fill="none" stroke="currentColor" viewBox="0 0 24 24">
      <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 13l4 4L19 7"/>
    </svg>
  </div>
  <h1 class="text-xl font-bold text-gray-900 mb-1">Importação concluída</h1>
  <p class="text-sm text-gray-500 mb-5">Dados da planilha carregados com sucesso.</p>
  <div class="bg-gray-50 border border-gray-200 rounded-xl p-4 text-left text-sm space-y-1 mb-5">
    <div class="flex justify-between"><span class="text-gray-500">Usuários criados</span><strong><?php echo $created; ?></strong></div>
    <div class="flex justify-between"><span class="text-gray-500">Usuários já existiam</span><strong><?php echo $skipped; ?></strong></div>
    <div class="flex justify-between"><span class="text-gray-500">Registros importados</span><strong><?php echo $recs; ?></strong></div>
  </div>
  <div class="bg-amber-50 border border-amber-200 rounded-xl p-3 text-xs text-amber-700 mb-5 text-left">
    <strong>Senha padrão dos colaboradores:</strong> Hostweb@2025<br>
    Eles precisarão redefinir no primeiro acesso.
  </div>
  <a href="/admin.php" class="block bg-blue-600 hover:bg-blue-700 text-white text-sm font-medium py-2.5 rounded-lg transition mb-3">
    Ir para o painel
  </a>
  <p class="text-xs text-red-500">⚠ Apague <code>import.php</code> do servidor após usar.</p>
</div>
</body></html>
""")

with open(OUTPUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

print(f"\n✓ Gerado: {OUTPUT}")
print("→ Acesse localhost:3000/import.php como admin para importar os dados.")
